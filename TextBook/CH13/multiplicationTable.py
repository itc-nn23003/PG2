import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def create_multiplication_table(n):
    # 新しいワークブックを作成する
    wb = Workbook()
    sheet = wb.active
    sheet.title = f"{n}x{n} Multiplication Table"

    # 最初の行と最初の列にラベルを入力する
    for i in range(1, n+1):
        sheet.cell(row=1, column=i+1, value=i)
        sheet.cell(row=i+1, column=1, value=i)

    # 掛け算の値を入力する
    for i in range(1, n+1):
        for j in range(1, n+1):
            sheet.cell(row=i+1, column=j+1, value=i*j)

    # 最初の行と列に太字のフォントと配置を設定します
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center')
    
    for row in sheet.iter_rows(min_row=1, max_row=1, min_col=2, max_col=n+1):
        for cell in row:
            cell.font = bold_font
            cell.alignment = center_alignment

    for col in sheet.iter_cols(min_row=2, max_row=n+1, min_col=1, max_col=1):
        for cell in col:
            cell.font = bold_font
            cell.alignment = center_alignment

    return wb

def main():
    if len(sys.argv) < 2:
        print("Usage: python script.py N")
        return
    
    try:
        N = int(sys.argv[1])
        if N <= 0:
            print("N must be a positive integer")
            return
    except ValueError:
        print("N must be a valid integer")
        return

    # 九九ワークブックを作成する
    wb = create_multiplication_table(N)

    # ワークブックをファイルに保存する
    output_file = f"multiplication_table_{N}x{N}.xlsx"
    wb.save(output_file)
    print(f"Multiplication table saved to {output_file}")

if __name__ == "__main__":
    main()

