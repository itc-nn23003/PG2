import openpyxl
import sys
import os

# コマンドライン引数が正しく与えられているかをチェックする
if len(sys.argv) != 2:
    print("Usage: python script.py <input_file>")
    sys.exit(1)

# 入力ファイルのパスを取得する
input_file = sys.argv[1]

# ファイルの存在を確認する
if not os.path.exists(input_file):
    print(f"Error: File '{input_file}' not found.")
    sys.exit(1)

# ワークブックを読み込む
try:
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
except Exception as e:
    print(f"Error: Failed to load '{input_file}': {e}")
    sys.exit(1)

# シートの行数と列数を取得する
max_row = sheet.max_row
max_col = sheet.max_column

# 新しいワークブックを作成し、入れ替えたデータを格納する
transposed_wb = openpyxl.Workbook()
transposed_sheet = transposed_wb.active
transposed_sheet.title = 'Transposed'

# データを入れ替える
for row in range(1, max_row + 1):
    for col in range(1, max_col + 1):
        # 元のシートからセルの値を読み出す
        cell_value = sheet.cell(row=row, column=col).value
        # 入れ替えた位置に新しいシートに書き込む
        transposed_sheet.cell(row=col, column=row).value = cell_value

# 出力ファイル名を生成する
output_file = os.path.splitext(input_file)[0] + '_swap.xlsx'

# 出力ファイルの存在を確認し、存在する場合は削除する
if os.path.exists(output_file):
    os.remove(output_file)

# 入れ替えたワークブックを新しいファイルに保存する
transposed_wb.save(output_file)
print(f"入れ替えたシートを {output_file} に保存しました。")

