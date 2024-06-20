import openpyxl

# Excelファイルを読み込む
input_excel = 'texts.xlsx'
wb = openpyxl.load_workbook(input_excel)
ws = wb.active

# データが入っている列の数を取得する
max_cols = ws.max_column

# 各列ごとに処理を行う
for col_index in range(1, max_cols + 1):
    # 各列に対応するテキストファイル名を準備する
    output_filename = f'texts.xlsx_{col_index}.txt'
    
    # テキストファイルを書き込みモードで開く
    with open(output_filename, 'w', encoding='utf-8') as file:
        # 各列のすべての行について処理を行う
        for row_index in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_index, column=col_index)
            cell_value = cell.value
            if cell_value is not None:
                file.write(str(cell_value) + '\n')  # セルの値をテキストファイルに書き込む。改行を追加する
    
    print(f'列 {col_index} のデータを {output_filename} に書き込みました')

print('テキストファイルの生成が完了しました。')

wb.close()

